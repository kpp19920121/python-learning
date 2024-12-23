from openpyxl import load_workbook
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.styles import PatternFill
import os

# 用于存储每个开发人员的上一个任务结束日期
developer_last_end = {}

# 排除周末的函数：根据给定的工作日数计算结束日期
def calculate_end_date(start_date, work_days):
    """
    根据开始日期和工时计算结束日期，排除周六和周日。
    :param start_date: 开始日期 (datetime.date)
    :param work_days: 任务所需的工作日数 (int)
    :return: 结束日期 (datetime.date)
    """
    current_date = start_date
    days_added = 0  # 已添加的工作日数量

    # 计算结束日期，排除周末
    while days_added < work_days - 1:  # 结束日期为开始日期加上工时减去1
        current_date += timedelta(days=1)

        # 如果是工作日（周一到周五），则增加工作日计数
        if current_date.weekday() < 5:  # weekday(): 0=Monday, 6=Sunday
            days_added += 1

    return current_date  # 返回计算后的结束日期

# 调整开始日期至下一个工作日（排除周六和周日）
def adjust_to_next_workday(start_date):
    """
    将开始日期调整为下一个工作日，排除周六和周日。
    :param start_date: 原始开始日期 (datetime.date)
    :return: 下一个工作日 (datetime.date)
    """
    while start_date.weekday() >= 5:  # 如果是周六（5）或周日（6），则跳到下一个工作日
        start_date += timedelta(days=1)
    return start_date

def calculate_DevloperLastEndDate(excel_file_path):



    print("开始获取excel中所有开发的最晚结束日期作为下一个计划的开始日期:")

    """
    计算每个开发人员的上一个任务结束日期。
    :param excel_file_path: Excel 文件路径 (str)
    :return: 一个字典，键为开发人员，值为对应的最后结束日期 (dict)
    """
    global developer_last_end

    developer_last_end={}
    print(f"这是测试的:{developer_last_end}")




    if  developer_last_end:
        print("字典不为空，直接返回")
        return developer_last_end
        # 空字典返回 True

    print("字典为空，开始查找开发人员的最晚计划结束日期")

    # 加载 Excel 文件
    wb = load_workbook(excel_file_path)
    sheet = wb.active

    # 将工作表数据加载为 DataFrame
    data = sheet.values
    columns = next(data)  # 读取第一行作为列名
    df = pd.DataFrame(data, columns=columns)

    # 检查是否包含必要的列
    required_columns = ["开发人员", "结束日期"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Excel 文件缺少必要列：{col}")

    # 存储开发人员的最后结束日期

    # 遍历每行数据，更新每个开发人员的最晚结束日期
    for idx, row in df.iterrows():
        developer = row["开发人员"]
        end_date = row["结束日期"]
        print(f"开发人员->{developer}:结束日期->{end_date}")

        # 如果结束日期为空，跳过该行
        if pd.isna(end_date):
            continue

        # 将结束日期转换为日期类型
        end_date = pd.to_datetime(end_date).date()

        print(f"end_date==============>{end_date}")

        print(f"end_date=>{end_date}:developer_last_end=>{developer_last_end}")

        # 如果该开发人员尚未记录结束日期，则直接赋值
        if developer not in developer_last_end:
            developer_last_end[developer] = end_date
        else:
            # 如果该开发人员已经有结束日期，选择较晚的日期
            if end_date > developer_last_end[developer]:
                developer_last_end[developer] = end_date


    for developer, last_end in developer_last_end.items():
        print(f"Developer: {developer}, Last End Date: {last_end}")

    return developer_last_end  # 返回开发人员的最晚结束日期字典

# 根据开发人员的排程情况计算开始日期
def calculate_start_date(developer, developer_last_end, work_days):

    """
    计算每个开发人员的任务开始日期。如果该开发人员已有任务，开始日期从上一个任务结束的下一工作日开始。
    如果是第一次安排任务，则从当前的第一个工作日开始。
    :param developer: 开发人员姓名 (str)
    :param developer_last_end: 开发人员上一个任务的结束日期 (dict)
    :param work_days: 任务所需的工作日数 (int)
    :return: 计算后的开始日期 (datetime.date)
    """
    # 如果开发人员已有排程，开始日期从上一个任务结束的下一工作日开始
    if developer in developer_last_end:
        start_date = developer_last_end[developer] + timedelta(days=1)
        start_date = adjust_to_next_workday(start_date)
    else:
        # 如果是第一次安排任务，从当前日期开始
        #start_date = adjust_to_next_workday(datetime.now().date())
        start_date = default_start_date


    return start_date

# 文件路径配置


# 用户输入 Excel 文件路径
input_file = input("请输入 Excel 文件路径（如未指定路径，将退出程序）：").strip()
if not input_file or not os.path.exists(input_file):
    raise FileNotFoundError(f"文件路径无效或不存在：{input_file}")

output_file = input("请输入输出文件路径：").strip()
if not output_file:
    output_file = os.path.splitext(input_file)[0] + "_排好的开发计划.xlsx"


# 用户输入默认开始日期
default_start_date_input = input("请输入默认开始日期（格式：YYYY-MM-DD，留空则为下一个工作日）：").strip()
if default_start_date_input:
    default_start_date = datetime.strptime(default_start_date_input, "%Y-%m-%d").date()
    default_start_date = adjust_to_next_workday(default_start_date)
else:
    default_start_date = adjust_to_next_workday(datetime.now().date())

print(f"这是测试的:{developer_last_end}")


calculate_DevloperLastEndDate(input_file)




# 加载 Excel 文件
wb = load_workbook(input_file)
sheet = wb.active

# 将工作表数据加载为 DataFrame
data = sheet.values
columns = next(data)  # 读取第一行作为列名
df = pd.DataFrame(data, columns=columns)

# 检查是否包含必要的列
required_columns = ["开发人员", "工时"]
for col in required_columns:
    if col not in df.columns:
        raise ValueError(f"Excel 文件缺少必要列：{col}")

# 初始化“开始日期”和“结束日期”列
df["开始日期"] = df.get("开始日期", None)  # 如果没有“开始日期”列则初始化为 None
df["结束日期"] = df.get("结束日期", None)  # 如果没有“结束日期”列则初始化为 None

# 定义红色填充样式
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
# 定义底纹样式（例如浅黄色底纹）
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 按行遍历并计算任务的开始和结束日期
for idx, row in df.iterrows():
    developer = row["开发人员"]
    work_days = row["工时"]
    start_date = row["开始日期"]
    end_date = row["结束日期"]

    #如果开发人员为空，则直接跳过
    if pd.isna(developer) or "/" in developer or  "-" in developer:
        continue

    # 如果开始日期或结束日期为空，则需要计算
    if pd.isna(start_date) or pd.isna(end_date):
        # 计算开始日期


        #如果指定了开始日期，则按照开始日期直接计算
        if pd.isna(start_date):
            start_date = calculate_start_date(developer, developer_last_end, work_days)

        # 计算结束日期
        end_date = calculate_end_date(start_date, work_days)

        # 确保开始日期和结束日期只包含年月日，去掉时分秒


        df.at[idx, "开始日期"] = pd.to_datetime(start_date)
        df.at[idx, "结束日期"] = pd.to_datetime(end_date)

        # 为修改过的单元格添加底纹
        #sheet.cell(row=idx + 2, column=columns.index("开始日期") + 1).fill = yellow_fill
        #heet.cell(row=idx + 2, column=columns.index("结束日期") + 1).fill = yellow_fill
        # 为修改过的整行添加红色底纹
        for col_idx in range(1, len(columns) + 1):  # 遍历整行
            sheet.cell(row=idx + 2, column=col_idx).fill = red_fill  # 为整行每个单元格添加红色底纹



        # 更新开发人员的最后结束日期
        developer_last_end[developer] = end_date

# 更新 Excel 工作表中的“开始日期”和“结束日期”
for r_idx, row in df.iterrows():
    # 将“开始时间”和“结束时间”写入对应单元格（仅写日期部分）
    sheet.cell(row=r_idx + 2, column=columns.index("开始日期") + 1, value=row["开始日期"])
    sheet.cell(row=r_idx + 2, column=columns.index("结束日期") + 1, value=row["结束日期"])

# 自动调整列宽（包括日期列）
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列名（如'A'、'B'、'C'等）

    for cell in col:
        try:
            # 对于日期格式的列，计算最大字符长度
            if isinstance(cell.value, datetime):
                # 获取日期字符串的长度（例如：2024-12-25）
                cell_length = len(str(cell.value.date()))
            else:
                cell_length = len(str(cell.value))

            if cell_length > max_length:
                max_length = cell_length
        except:
            pass

    adjusted_width = (max_length + 2)  # 添加2个字符的间距
    sheet.column_dimensions[column].width = adjusted_width

# 保存结果
wb.save(output_file)
print(f"开发计划已经排好，并保存到文件：{output_file}")
